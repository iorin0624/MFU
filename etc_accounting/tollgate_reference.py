from __future__ import annotations

import hashlib
import os
import re
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

import requests
from openpyxl import load_workbook

from app.utils.db import get_db


SOURCE_URL = os.environ.get(
    "MFU_ETC_TOLLGATE_SOURCE_URL",
    "https://www.etc-meisai.jp/faq/files/Tollgatesearch.xlsx",
)
REFERENCE_ROOT = Path(
    os.environ.get("MFU_ETC_TOLLGATE_ROOT", "/mnt/mfu/etc_reference")
)
REFERENCE_PATH = REFERENCE_ROOT / "Tollgatesearch.xlsx"
REFERENCE_SHEET = "（参考）料金所一覧"
MINIMUM_REFERENCE_ROWS = 1000
DOWNLOAD_LIMIT_BYTES = 20 * 1024 * 1024


def normalize_tollgate_name(value: object) -> str:
    text = unicodedata.normalize("NFKC", str(value or ""))
    return re.sub(r"\s+", "", text).strip()


def _reference_lookup(rows: Iterable[dict]) -> dict[str, list[tuple[str, str, str]]]:
    grouped: dict[str, set[tuple[str, str, str]]] = defaultdict(set)
    for row in rows:
        normalized = normalize_tollgate_name(
            row.get("normalized_name") or row.get("tollgate_name")
        )
        operator_name = str(row.get("operator_name") or "").strip()
        road_name = str(row.get("road_name") or "").strip()
        tollgate_name = str(row.get("tollgate_name") or "").strip()
        if normalized and operator_name and road_name and tollgate_name:
            grouped[normalized].add((operator_name, road_name, tollgate_name))
    return {
        normalized: sorted(values)
        for normalized, values in grouped.items()
    }


def resolve_exit_tollgate(exit_ic: object, lookup: dict) -> dict:
    normalized = normalize_tollgate_name(exit_ic)
    if not normalized:
        return {
            "status": "no_exit",
            "operator_name": None,
            "road_name": None,
            "matched_name": None,
        }
    matches = list(lookup.get(normalized) or [])
    mappings = {(operator_name, road_name) for operator_name, road_name, _name in matches}
    if not matches:
        return {
            "status": "unmatched",
            "operator_name": None,
            "road_name": None,
            "matched_name": None,
        }
    if len(mappings) != 1:
        return {
            "status": "ambiguous",
            "operator_name": None,
            "road_name": None,
            "matched_name": None,
        }
    operator_name, road_name = next(iter(mappings))
    matched_names = sorted({name for _operator, _road, name in matches})
    return {
        "status": "matched",
        "operator_name": operator_name,
        "road_name": road_name,
        "matched_name": matched_names[0],
    }


def parse_reference_workbook(
    path: Path,
    *,
    minimum_rows: int = MINIMUM_REFERENCE_ROWS,
) -> list[dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if REFERENCE_SHEET not in workbook.sheetnames:
            raise RuntimeError(f"料金所一覧シートがありません: {REFERENCE_SHEET}")
        sheet = workbook[REFERENCE_SHEET]
        header = [sheet.cell(5, column).value for column in range(1, 6)]
        if header[0] != "会社・公社" or header[3] != "料金所等名":
            raise RuntimeError("料金所一覧の列構成が想定と異なります。")
        rows = []
        for source_row, values in enumerate(
            sheet.iter_rows(min_row=6, max_col=5, values_only=True),
            start=6,
        ):
            operator_name = str(values[0] or "").strip()
            road_name = str(values[1] or "").strip()
            tollgate_name = str(values[3] or "").strip()
            tollgate_reading = str(values[4] or "").strip()
            normalized_name = normalize_tollgate_name(tollgate_name)
            if not operator_name or not road_name or not normalized_name:
                continue
            rows.append(
                {
                    "operator_name": operator_name,
                    "road_name": road_name,
                    "tollgate_name": tollgate_name,
                    "tollgate_reading": tollgate_reading or None,
                    "normalized_name": normalized_name,
                    "source_row": source_row,
                }
            )
        if len(rows) < int(minimum_rows):
            raise RuntimeError(
                f"料金所一覧の件数が少なすぎます: {len(rows)}件"
            )
        return rows
    finally:
        workbook.close()


def _ensure_schema() -> None:
    from .repository import ensure_schema

    ensure_schema()


def _save_reference_state(
    *,
    status: str,
    row_count: int = 0,
    sha256: str | None = None,
    etag: str | None = None,
    last_modified: str | None = None,
    error: str | None = None,
) -> None:
    _ensure_schema()
    now = datetime.now()
    db = get_db()
    try:
        cur = db.cursor()
        cur.execute(
            """
            INSERT INTO etc_tollgate_reference_state (
                id, source_url, source_sha256, source_etag, source_last_modified,
                row_count, status, error_text, checked_at, updated_at
            ) VALUES (1,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON DUPLICATE KEY UPDATE
                source_url=VALUES(source_url),
                source_sha256=COALESCE(VALUES(source_sha256), source_sha256),
                source_etag=COALESCE(VALUES(source_etag), source_etag),
                source_last_modified=COALESCE(VALUES(source_last_modified), source_last_modified),
                row_count=IF(VALUES(status)='success', VALUES(row_count), row_count),
                status=VALUES(status), error_text=VALUES(error_text),
                checked_at=VALUES(checked_at), updated_at=VALUES(updated_at)
            """,
            (
                SOURCE_URL,
                sha256,
                etag,
                last_modified,
                int(row_count),
                status[:32],
                (error or "")[:4000] or None,
                now,
                now,
            ),
        )
        db.commit()
    finally:
        db.close()


def _replace_reference_rows(rows: list[dict]) -> None:
    _ensure_schema()
    now = datetime.now()
    values = [
        (
            row["operator_name"],
            row["road_name"],
            row["tollgate_name"],
            row.get("tollgate_reading"),
            row["normalized_name"],
            int(row["source_row"]),
            now,
            now,
        )
        for row in rows
    ]
    db = get_db()
    try:
        cur = db.cursor()
        db.start_transaction()
        cur.execute("DELETE FROM etc_tollgate_reference")
        for offset in range(0, len(values), 500):
            cur.executemany(
                """
                INSERT INTO etc_tollgate_reference (
                    operator_name, road_name, tollgate_name, tollgate_reading,
                    normalized_name, source_row, created_at, updated_at
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                """,
                values[offset : offset + 500],
            )
        db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def _load_reference_rows() -> list[dict]:
    _ensure_schema()
    db = get_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute(
            """
            SELECT operator_name, road_name, tollgate_name, normalized_name
            FROM etc_tollgate_reference
            """
        )
        return cur.fetchall()
    finally:
        db.close()


def enrich_record_tollgate(record_id: int, exit_ic: object) -> dict:
    lookup = _reference_lookup(_load_reference_rows())
    if not lookup:
        return {"status": "reference_unavailable"}
    match = resolve_exit_tollgate(exit_ic, lookup)
    now = datetime.now()
    db = get_db()
    try:
        cur = db.cursor()
        cur.execute(
            """
            UPDATE etc_freee_records
            SET tollgate_operator_name=%s, tollgate_road_name=%s,
                tollgate_matched_name=%s, tollgate_match_status=%s,
                tollgate_reference_updated_at=%s
            WHERE id=%s
            """,
            (
                match["operator_name"],
                match["road_name"],
                match["matched_name"],
                match["status"],
                now,
                int(record_id),
            ),
        )
        db.commit()
    finally:
        db.close()
    return match


def backfill_record_tollgates() -> dict:
    reference_rows = _load_reference_rows()
    lookup = _reference_lookup(reference_rows)
    if not lookup:
        return {"updated": 0, "matched": 0, "unmatched": 0, "ambiguous": 0}
    db = get_db()
    try:
        cur = db.cursor(dictionary=True)
        cur.execute("SELECT id, exit_ic FROM etc_freee_records ORDER BY id")
        records = cur.fetchall()
        now = datetime.now()
        results = []
        counts = defaultdict(int)
        for record in records:
            match = resolve_exit_tollgate(record.get("exit_ic"), lookup)
            counts[match["status"]] += 1
            results.append(
                (
                    match["operator_name"],
                    match["road_name"],
                    match["matched_name"],
                    match["status"],
                    now,
                    int(record["id"]),
                )
            )
        if results:
            cur.executemany(
                """
                UPDATE etc_freee_records
                SET tollgate_operator_name=%s, tollgate_road_name=%s,
                    tollgate_matched_name=%s, tollgate_match_status=%s,
                    tollgate_reference_updated_at=%s
                WHERE id=%s
                """,
                results,
            )
        db.commit()
        return {
            "updated": len(results),
            "matched": int(counts["matched"]),
            "unmatched": int(counts["unmatched"] + counts["no_exit"]),
            "ambiguous": int(counts["ambiguous"]),
        }
    finally:
        db.close()


def _download_reference(target: Path) -> dict:
    response = requests.get(
        SOURCE_URL,
        headers={"User-Agent": "MFU ETC tollgate reference updater/1.0"},
        stream=True,
        timeout=(10, 60),
    )
    response.raise_for_status()
    size = 0
    digest = hashlib.sha256()
    with target.open("wb") as output:
        for chunk in response.iter_content(chunk_size=64 * 1024):
            if not chunk:
                continue
            size += len(chunk)
            if size > DOWNLOAD_LIMIT_BYTES:
                raise RuntimeError("料金所一覧ファイルが上限サイズを超えました。")
            digest.update(chunk)
            output.write(chunk)
    if size < 10_000:
        raise RuntimeError("料金所一覧ファイルが小さすぎます。")
    with target.open("rb") as downloaded:
        if downloaded.read(2) != b"PK":
            raise RuntimeError("取得結果がExcelファイルではありません。")
    return {
        "sha256": digest.hexdigest(),
        "etag": response.headers.get("ETag"),
        "last_modified": response.headers.get("Last-Modified"),
    }


def refresh_tollgate_reference() -> dict:
    REFERENCE_ROOT.mkdir(parents=True, exist_ok=True)
    try:
        os.chmod(REFERENCE_ROOT, 0o700)
    except OSError:
        pass
    temporary = REFERENCE_ROOT / "Tollgatesearch.part.xlsx"
    temporary.unlink(missing_ok=True)
    metadata: dict = {}
    try:
        metadata = _download_reference(temporary)
        rows = parse_reference_workbook(temporary)
        os.chmod(temporary, 0o600)
        temporary.replace(REFERENCE_PATH)
        _replace_reference_rows(rows)
        backfill = backfill_record_tollgates()
        _save_reference_state(
            status="success",
            row_count=len(rows),
            sha256=metadata["sha256"],
            etag=metadata.get("etag"),
            last_modified=metadata.get("last_modified"),
        )
        return {
            "status": "success",
            "row_count": len(rows),
            "sha256": metadata["sha256"],
            "backfill": backfill,
        }
    except Exception as exc:
        _save_reference_state(
            status="error",
            etag=metadata.get("etag"),
            last_modified=metadata.get("last_modified"),
            error=str(exc),
        )
        raise
    finally:
        temporary.unlink(missing_ok=True)
