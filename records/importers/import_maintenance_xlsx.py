import sys
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from app.utils.db import get_db
from app.records.models import ensure_maintenance_schema


def _parse_excel_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError:
            try:
                return datetime.strptime(text, "%Y/%m/%d").date()
            except ValueError:
                return None
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except ValueError:
            return None
    return None


def import_maintenance_xlsx(path: str) -> int:
    wb = load_workbook(path, data_only=True)
    if "整備一覧" not in wb.sheetnames:
        raise ValueError("シート『整備一覧』が見つかりません")
    ws = wb["整備一覧"]

    db = get_db()
    ensure_maintenance_schema(db)
    cur = db.cursor()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        event_date = _parse_excel_date(row[0])
        if event_date is None:
            continue
        odometer_km = int(row[1] or 0)
        item = str(row[2] or "").strip()
        note = str(row[3] or "").strip() or None
        if not item:
            continue
        now = datetime.now()
        cur.execute(
            """
            INSERT INTO bike_maintenance_log (
                event_date,
                odometer_km,
                item,
                note,
                created_at,
                updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (event_date, odometer_km, item, note, now, now),
        )
        count += 1

    db.commit()
    db.close()
    return count


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python -m app.records.importers.import_maintenance_xlsx /path/to/バイク整備記録.xlsx")
    path = sys.argv[1]
    total = import_maintenance_xlsx(path)
    print(f"imported {total} rows")


if __name__ == "__main__":
    main()
