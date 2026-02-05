import sys
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from app.utils.db import get_db
from app.records.models import ensure_uber_schema


def _parse_excel_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            return datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError:
            try:
                return datetime.strptime(text, "%Y/%m/%d").date()
            except ValueError:
                return None
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date()
        except ValueError:
            return None
    return None


def import_uber_xlsx(path: str) -> int:
    wb = load_workbook(path, data_only=True)
    if "一覧" not in wb.sheetnames:
        raise ValueError("シート『一覧』が見つかりません")
    ws = wb["一覧"]

    db = get_db()
    ensure_uber_schema(db)
    cur = db.cursor()

    count = 0
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        work_date = _parse_excel_date(row[0])
        if work_date is None:
            continue
        deliveries = int(row[1] or 0)
        net_yen = int(row[2] or 0)
        promo_yen = int(row[3] or 0)
        other_yen = int(row[4] or 0)
        tip_yen = int(row[5] or 0)
        now = datetime.now()
        cur.execute(
            """
            INSERT INTO uber_daily (
                work_date,
                deliveries,
                net_yen,
                promo_yen,
                other_yen,
                tip_yen,
                created_at,
                updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                deliveries = VALUES(deliveries),
                net_yen = VALUES(net_yen),
                promo_yen = VALUES(promo_yen),
                other_yen = VALUES(other_yen),
                tip_yen = VALUES(tip_yen),
                updated_at = VALUES(updated_at)
            """,
            (
                work_date,
                deliveries,
                net_yen,
                promo_yen,
                other_yen,
                tip_yen,
                now,
                now,
            ),
        )
        count += 1

    db.commit()
    db.close()
    return count


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python -m app.records.importers.import_uber_xlsx /path/to/Uber記録.xlsx")
    path = sys.argv[1]
    total = import_uber_xlsx(path)
    print(f"imported {total} rows")


if __name__ == "__main__":
    main()
